import os
import json
import datetime
import hashlib
import secrets
import threading
import io
import re
import time
import shutil
import math
from typing import Optional

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from fastapi import FastAPI, HTTPException, Header, Depends, Query, Form, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.requests import Request
from pydantic import BaseModel, Field

from backend.logger import setup_logger
from backend.json_storage import read_json, write_json, update_json, BACKUPS_DIR
from backend.security import sanitize_html, sanitize_dict, validate_email, validate_phone, hash_password, verify_password
from backend.exceptions import ErrorHandlingMiddleware, register_exception_handlers

logger = setup_logger("solar_app")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STOCK_FILE_PATH = os.environ.get("STOCK_FILE_PATH") or os.path.join(BASE_DIR, "data", "stock.json")
REPORTS_DIR = os.environ.get("REPORTS_DIR") or os.path.join(BASE_DIR, "reports")
IMAGES_DIR = os.environ.get("IMAGES_DIR") or os.path.join(BASE_DIR, "data", "images")
LOGS_DIR = os.environ.get("LOGS_DIR") or os.path.join(BASE_DIR, "logs")

os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

active_tokens = {}
TOKEN_EXPIRY_SECONDS = 86400 * 7

app = FastAPI(
    title="Solar Stock Management API",
    version="2.0.0",
    docs_url=None,
    redoc_url=None
)

allowed_origins = os.environ.get("CORS_ORIGINS", "*")
if allowed_origins == "*":
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    origins = [o.strip() for o in allowed_origins.split(",")]
    app.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

app.add_middleware(ErrorHandlingMiddleware)
register_exception_handlers(app)

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; img-src 'self' data: https://images.unsplash.com; font-src 'self' https://cdn.jsdelivr.net https://fonts.gstatic.com; connect-src 'self' https://cdn.jsdelivr.net"
    return response

@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    elapsed = time.time() - start
    logger.info(f"{request.method} {request.url.path} -> {response.status_code} ({elapsed:.3f}s)")
    return response

try:
    app.mount("/api/images", StaticFiles(directory=IMAGES_DIR), name="images")
except Exception as e:
    logger.warning(f"Could not mount /api/images: {e}")

from backend.scheduler import init_scheduler
scheduler = init_scheduler(STOCK_FILE_PATH)

@app.on_event("startup")
async def startup_event():
    logger.info("Solar Stock Management Backend v2.0.0 starting...")
    try:
        scheduler.start()
        logger.info("Background scheduler started successfully.")
    except Exception as e:
        logger.error(f"Failed to start scheduler: {e}")

@app.on_event("shutdown")
async def shutdown_event():
    logger.info("Shutting down Solar Stock Management Backend...")
    try:
        if scheduler and scheduler.running:
            scheduler.shutdown()
            logger.info("Scheduler shut down.")
    except Exception as e:
        logger.error(f"Scheduler shutdown error: {e}")

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
MAX_IMAGE_SIZE = 5 * 1024 * 1024

class LoginRequest(BaseModel):
    username: str = Field(..., min_length=1, max_length=100)
    password: str = Field(..., min_length=1, max_length=200)

class ProductModel(BaseModel):
    id: str = Field(..., min_length=1, max_length=50, pattern=r'^[A-Za-z0-9_\-]+$')
    name: str = Field(..., min_length=1, max_length=200)
    category: str = Field(..., max_length=100)
    brand: str = Field(..., max_length=100)
    unit: str = Field(..., max_length=50)
    quantity: int = Field(ge=0)
    minimum_stock: int = Field(ge=0)
    supplier: str = Field(default="", max_length=200)
    rack_location: str = Field(default="", max_length=200)
    model_capacity: Optional[str] = Field(default="", max_length=200)
    image_path: Optional[str] = Field(default="", max_length=500)

class SupplierModel(BaseModel):
    id: str = Field(..., min_length=1, max_length=50, pattern=r'^[A-Za-z0-9_\-]+$')
    name: str = Field(..., min_length=1, max_length=200)
    contact_person: Optional[str] = Field(default="", max_length=200)
    phone: Optional[str] = Field(default="", max_length=50)
    email: Optional[str] = Field(default="", max_length=200)
    address: Optional[str] = Field(default="", max_length=500)

class CustomerModel(BaseModel):
    id: str = Field(..., min_length=1, max_length=50, pattern=r'^[A-Za-z0-9_\-]+$')
    name: str = Field(..., min_length=1, max_length=200)
    contact_person: Optional[str] = Field(default="", max_length=200)
    phone: Optional[str] = Field(default="", max_length=50)
    email: Optional[str] = Field(default="", max_length=200)
    address: Optional[str] = Field(default="", max_length=500)

class StockInRequest(BaseModel):
    product_id: str = Field(..., min_length=1, max_length=50)
    quantity: int = Field(gt=0)
    supplier: Optional[str] = Field(default="", max_length=200)
    date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    remarks: Optional[str] = Field(default="", max_length=500)

class StockOutRequest(BaseModel):
    product_id: str = Field(..., min_length=1, max_length=50)
    quantity: int = Field(gt=0)
    customer: Optional[str] = Field(default="", max_length=200)
    employee: Optional[str] = Field(default="", max_length=200)
    date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    remarks: Optional[str] = Field(default="", max_length=500)

class EmailConfigModel(BaseModel):
    email_smtp_server: str = Field(..., max_length=200)
    email_smtp_port: int = Field(default=587, ge=1, le=65535)
    email_sender: str = Field(..., max_length=200)
    email_password: str = Field(..., max_length=500)
    email_recipient: str = Field(..., max_length=200)

class WhatsAppConfigModel(BaseModel):
    whatsapp_recipient: str
    whatsapp_phone_number_id: str
    whatsapp_token: str

def verify_token(authorization: Optional[str] = Header(None)) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authentication token missing or invalid format")
    token = authorization.split(" ")[1]
    token_data = active_tokens.get(token)
    if token_data is None:
        raise HTTPException(status_code=401, detail="Session expired or invalid token")
    created_at = token_data.get("created_at", 0)
    if time.time() - created_at > TOKEN_EXPIRY_SECONDS:
        active_tokens.pop(token, None)
        raise HTTPException(status_code=401, detail="Session expired. Please log in again.")
    return token

def verify_token_optional(authorization: Optional[str] = Header(None)) -> Optional[str]:
    if not authorization or not authorization.startswith("Bearer "):
        return None
    token = authorization.split(" ")[1]
    token_data = active_tokens.get(token)
    if token_data is None:
        return None
    created_at = token_data.get("created_at", 0)
    if time.time() - created_at > TOKEN_EXPIRY_SECONDS:
        active_tokens.pop(token, None)
        return None
    return token

@app.get("/api/health")
def health_check():
    return {
        "status": "healthy",
        "version": "2.0.0",
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "uptime": time.time() - startup_time
    }

startup_time = time.time()

@app.post("/api/login")
def login(payload: LoginRequest):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})

    username = payload.username.strip()
    password = payload.password

    salt = config.get("admin_salt", "")
    stored_hash = config.get("admin_password_hash", "")

    if verify_password(password, salt, stored_hash) and username == config.get("admin_username"):
        token = secrets.token_hex(32)
        active_tokens[token] = {"created_at": time.time(), "username": username}
        logger.info(f"User '{username}' logged in successfully")
        return {"success": True, "token": token, "username": username}
    else:
        logger.warning(f"Failed login attempt for user '{username}'")
        raise HTTPException(status_code=401, detail="Invalid username or password")

@app.post("/api/logout")
def logout(token: str = Depends(verify_token)):
    active_tokens.pop(token, None)
    logger.info("User logged out")
    return {"success": True}

@app.get("/api/dashboard")
def get_dashboard_data(request: Request, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    transactions = data.get("transactions", [])

    total_products = len(products)
    available_stock = sum(p.get("quantity", 0) for p in products)

    low_stock_items = 0
    out_of_stock_items = 0
    for p in products:
        q = p.get("quantity", 0)
        min_s = p.get("minimum_stock", 0)
        if q == 0:
            out_of_stock_items += 1
        if q > 0 and q < min_s:
            low_stock_items += 1

    today_iso = datetime.date.today().isoformat()
    today_stock_in = 0
    today_stock_out = 0

    for tx in transactions:
        tx_date = tx.get("date", "")
        if tx_date == today_iso:
            qty = tx.get("quantity", 0)
            if tx.get("type") == "IN":
                today_stock_in += qty
            elif tx.get("type") == "OUT":
                today_stock_out += qty

    categories_mix = {}
    for p in products:
        cat = p.get("category", "Others")
        categories_mix[cat] = categories_mix.get(cat, 0) + p.get("quantity", 0)

    last_7_days = []
    for i in range(6, -1, -1):
        d = (datetime.date.today() - datetime.timedelta(days=i)).isoformat()
        last_7_days.append(d)

    chart_in_out = {
        "labels": [datetime.datetime.strptime(d, "%Y-%m-%d").strftime("%d %b") for d in last_7_days],
        "in": [0] * 7,
        "out": [0] * 7
    }

    for tx in transactions:
        tx_date = tx.get("date", "")
        if tx_date in last_7_days:
            idx = last_7_days.index(tx_date)
            qty = tx.get("quantity", 0)
            if tx.get("type") == "IN":
                chart_in_out["in"][idx] += qty
            elif tx.get("type") == "OUT":
                chart_in_out["out"][idx] += qty

    stock_status_counts = {
        "Normal Stock": total_products - (low_stock_items + out_of_stock_items),
        "Low Stock": low_stock_items,
        "Out of Stock": out_of_stock_items
    }

    # Energy forecast — per panel type only
    config = data.get("config", {})
    temp_c = float(request.query_params.get("temp_c", config.get("energy_temp_celsius", 25)))
    peak_sun_hours = float(config.get("energy_peak_sun_hours", 5))
    temp_coeff = 1 - max(0, (temp_c - 25) * 0.004)

    panel_types = []
    for p in products:
        if p.get("category", "").lower() != "solar panels":
            continue
        if p.get("quantity", 0) < 1:
            continue
        w = 0
        cap = (p.get("model_capacity") or "").strip()
        if cap:
            m = re.search(r'(\d+(?:\.\d+)?)\s*k?[wW]', cap)
            if m:
                val = float(m.group(1))
                w = val * 1000 if 'k' in cap[0:cap.lower().index('w')+1].lower() else val
        if w == 0:
            m = re.search(r'(\d+(?:\.\d+)?)\s*[wW]', p.get("name", ""))
            if m:
                w = float(m.group(1))
        unit_w = w
        unit_daily_kwh = (unit_w / 1000) * peak_sun_hours * temp_coeff
        eff_pct = round(min(unit_w / 17.0, 24.0), 1) if unit_w else 18.0
        panel_types.append({
            "name": p.get("name", ""),
            "unit_watt": round(unit_w),
            "unit_daily_kwh": round(unit_daily_kwh, 3),
            "quantity": p.get("quantity", 0),
            "brand": p.get("brand", ""),
            "efficiency_pct": eff_pct
        })

    return {
        "metrics": {
            "total_products": total_products,
            "available_stock": available_stock,
            "low_stock_items": low_stock_items,
            "out_of_stock_items": out_of_stock_items,
            "today_stock_in": today_stock_in,
            "today_stock_out": today_stock_out
        },
        "charts": {
            "categories": categories_mix,
            "in_out_trend": chart_in_out,
            "stock_status": stock_status_counts
        },
        "energy_forecast": {
            "panel_types": panel_types,
            "temperature_c": temp_c,
            "peak_sun_hours": peak_sun_hours,
            "temperature_derating": round(temp_coeff, 3)
        }
    }

@app.get("/api/products", response_model=list)
def get_products(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    return data.get("products", [])

@app.post("/api/products/{product_id}/image")
async def upload_product_image(product_id: str, file: UploadFile = File(...), token: str = Depends(verify_token)):
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid file type '{file.content_type}'. Allowed: JPEG, PNG, WebP, GIF")
    contents = await file.read()
    if len(contents) > MAX_IMAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large ({len(contents)} bytes). Maximum: 5MB")
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    found_idx = -1
    for idx, p in enumerate(products):
        if p.get("id") == product_id:
            found_idx = idx
            break
    if found_idx == -1:
        raise HTTPException(status_code=404, detail="Product not found")
    ext_map = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}
    extension = ext_map.get(file.content_type, "jpg")
    safe_filename = f"{sanitize_html(product_id)}_{secrets.token_hex(4)}.{extension}"
    file_location = os.path.join(IMAGES_DIR, safe_filename)
    with open(file_location, "wb") as f:
        f.write(contents)
    image_url = f"/api/images/{safe_filename}"
    products[found_idx]["image_path"] = image_url
    data["products"] = products
    write_json(STOCK_FILE_PATH, data)
    logger.info(f"Image uploaded for product {product_id}: {image_url}")
    return {"success": True, "message": "Image uploaded successfully", "image_url": image_url}

@app.post("/api/products")
def add_product(product: ProductModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    if any(p.get("id") == product.id for p in products):
        raise HTTPException(status_code=400, detail=f"Product with ID {product.id} already exists")
    safe = sanitize_dict(product.dict())
    products.append(safe)
    data["products"] = products
    write_json(STOCK_FILE_PATH, data)
    logger.info(f"Product added: {product.id} - {product.name}")
    return {"success": True, "message": "Product added successfully", "product": safe}

@app.put("/api/products/{product_id}")
def edit_product(product_id: str, product: ProductModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    found_idx = -1
    for idx, p in enumerate(products):
        if p.get("id") == product_id:
            found_idx = idx
            break
    if found_idx == -1:
        raise HTTPException(status_code=404, detail="Product not found")
    product_dict = sanitize_dict(product.dict())
    product_dict["id"] = product_id
    products[found_idx] = product_dict
    data["products"] = products
    write_json(STOCK_FILE_PATH, data)
    logger.info(f"Product updated: {product_id}")
    return {"success": True, "message": "Product updated successfully", "product": product_dict}

@app.delete("/api/products/{product_id}")
def delete_product(product_id: str, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    initial_len = len(products)
    products = [p for p in products if p.get("id") != product_id]
    if len(products) == initial_len:
        raise HTTPException(status_code=404, detail="Product not found")
    data["products"] = products
    write_json(STOCK_FILE_PATH, data)
    logger.info(f"Product deleted: {product_id}")
    return {"success": True, "message": "Product deleted successfully"}

@app.get("/api/suppliers", response_model=list)
def get_suppliers(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    return data.get("suppliers", [])

@app.post("/api/suppliers")
def add_supplier(supplier: SupplierModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    suppliers = data.get("suppliers", [])
    if any(s.get("id") == supplier.id for s in suppliers):
        raise HTTPException(status_code=400, detail=f"Supplier with ID {supplier.id} already exists")
    if not validate_phone(supplier.phone):
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    safe = sanitize_dict(supplier.dict())
    suppliers.append(safe)
    data["suppliers"] = suppliers
    write_json(STOCK_FILE_PATH, data)
    return {"success": True, "message": "Supplier added", "supplier": safe}

@app.put("/api/suppliers/{supplier_id}")
def edit_supplier(supplier_id: str, supplier: SupplierModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    suppliers = data.get("suppliers", [])
    for idx, s in enumerate(suppliers):
        if s.get("id") == supplier_id:
            sup_dict = sanitize_dict(supplier.dict())
            sup_dict["id"] = supplier_id
            suppliers[idx] = sup_dict
            data["suppliers"] = suppliers
            write_json(STOCK_FILE_PATH, data)
            return {"success": True, "message": "Supplier updated", "supplier": sup_dict}
    raise HTTPException(status_code=404, detail="Supplier not found")

@app.delete("/api/suppliers/{supplier_id}")
def delete_supplier(supplier_id: str, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    suppliers = data.get("suppliers", [])
    initial_len = len(suppliers)
    suppliers = [s for s in suppliers if s.get("id") != supplier_id]
    if len(suppliers) == initial_len:
        raise HTTPException(status_code=404, detail="Supplier not found")
    data["suppliers"] = suppliers
    write_json(STOCK_FILE_PATH, data)
    return {"success": True, "message": "Supplier deleted"}

@app.get("/api/customers", response_model=list)
def get_customers(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    return data.get("customers", [])

@app.post("/api/customers")
def add_customer(customer: CustomerModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    customers = data.get("customers", [])
    if any(c.get("id") == customer.id for c in customers):
        raise HTTPException(status_code=400, detail=f"Customer with ID {customer.id} already exists")
    if not validate_phone(customer.phone):
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    safe = sanitize_dict(customer.dict())
    customers.append(safe)
    data["customers"] = customers
    write_json(STOCK_FILE_PATH, data)
    return {"success": True, "message": "Customer added", "customer": safe}

@app.put("/api/customers/{customer_id}")
def edit_customer(customer_id: str, customer: CustomerModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    customers = data.get("customers", [])
    for idx, c in enumerate(customers):
        if c.get("id") == customer_id:
            cust_dict = sanitize_dict(customer.dict())
            cust_dict["id"] = customer_id
            customers[idx] = cust_dict
            data["customers"] = customers
            write_json(STOCK_FILE_PATH, data)
            return {"success": True, "message": "Customer updated", "customer": cust_dict}
    raise HTTPException(status_code=404, detail="Customer not found")

@app.delete("/api/customers/{customer_id}")
def delete_customer(customer_id: str, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    customers = data.get("customers", [])
    initial_len = len(customers)
    customers = [c for c in customers if c.get("id") != customer_id]
    if len(customers) == initial_len:
        raise HTTPException(status_code=404, detail="Customer not found")
    data["customers"] = customers
    write_json(STOCK_FILE_PATH, data)
    return {"success": True, "message": "Customer deleted"}

@app.get("/api/transactions")
def get_transactions(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    txs = data.get("transactions", [])
    return sorted(txs, key=lambda x: x.get("timestamp", ""), reverse=True)

@app.post("/api/stock-in")
def stock_in(payload: StockInRequest, token: str = Depends(verify_token)):
    result = {"transaction": None}
    def updater(data):
        products = data.get("products", [])
        transactions = data.get("transactions", [])
        product = next((p for p in products if p.get("id") == payload.product_id), None)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        product["quantity"] = product.get("quantity", 0) + payload.quantity
        tx_num = 1
        if transactions:
            tx_ids = [tx.get("id", "T000") for tx in transactions]
            numeric_ids = [int(tid[1:]) for tid in tx_ids if tid.startswith("T") and tid[1:].isdigit()]
            if numeric_ids:
                tx_num = max(numeric_ids) + 1
        tx_id = f"T{tx_num:03d}"
        new_tx = {
            "id": tx_id,
            "type": "IN",
            "product_id": payload.product_id,
            "product_name": product.get("name", "Unknown"),
            "quantity": payload.quantity,
            "entity": payload.supplier,
            "date": payload.date,
            "remarks": payload.remarks,
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat().replace('+00:00', 'Z')
        }
        transactions.append(new_tx)
        data["products"] = products
        data["transactions"] = transactions
        result["transaction"] = new_tx
        logger.info(f"Stock IN: {payload.quantity} x {product.get('name')} (TX: {tx_id})")
        return data
    update_json(STOCK_FILE_PATH, updater)
    return {"success": True, "message": f"Added {payload.quantity} items", "transaction": result["transaction"]}

@app.post("/api/stock-out")
def stock_out(payload: StockOutRequest, token: str = Depends(verify_token)):
    result = {"transaction": None}
    def updater(data):
        products = data.get("products", [])
        transactions = data.get("transactions", [])
        product = next((p for p in products if p.get("id") == payload.product_id), None)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        old_qty = product.get("quantity", 0)
        if old_qty < payload.quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock. Available: {old_qty}")
        product["quantity"] = old_qty - payload.quantity
        tx_num = 1
        if transactions:
            tx_ids = [tx.get("id", "T000") for tx in transactions]
            numeric_ids = [int(tid[1:]) for tid in tx_ids if tid.startswith("T") and tid[1:].isdigit()]
            if numeric_ids:
                tx_num = max(numeric_ids) + 1
        tx_id = f"T{tx_num:03d}"
        new_tx = {
            "id": tx_id,
            "type": "OUT",
            "product_id": payload.product_id,
            "product_name": product.get("name", "Unknown"),
            "quantity": payload.quantity,
            "entity": payload.customer or "",
            "employee": payload.employee or "",
            "date": payload.date,
            "remarks": payload.remarks,
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat().replace('+00:00', 'Z')
        }
        transactions.append(new_tx)
        data["products"] = products
        data["transactions"] = transactions
        result["transaction"] = new_tx
        logger.info(f"Stock OUT: {payload.quantity} x {product.get('name')} (TX: {tx_id})")
        return data
    update_json(STOCK_FILE_PATH, updater)
    return {"success": True, "message": f"Deducted {payload.quantity} items", "transaction": result["transaction"]}

@app.get("/api/email/config")
def get_email_config(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    pw = config.get("email_password", "")
    masked = "***" + pw[-4:] if pw and len(pw) > 4 else "***" if pw else ""
    return {
        "email_smtp_server": config.get("email_smtp_server", ""),
        "email_smtp_port": config.get("email_smtp_port", 587),
        "email_sender": config.get("email_sender", ""),
        "email_password": masked,
        "email_recipient": config.get("email_recipient", "")
    }

@app.post("/api/email/config")
def save_email_config(payload: EmailConfigModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    config["email_smtp_server"] = payload.email_smtp_server.strip()
    config["email_smtp_port"] = payload.email_smtp_port
    config["email_sender"] = payload.email_sender.strip()
    if payload.email_password and not payload.email_password.startswith("***"):
        config["email_password"] = payload.email_password
    config["email_recipient"] = payload.email_recipient.strip()
    data["config"] = config
    write_json(STOCK_FILE_PATH, data)
    logger.info("Email configuration updated")
    return {"success": True, "message": "Email settings updated successfully"}

@app.post("/api/email/test-send")
def trigger_email_test_send(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    transactions = data.get("transactions", [])
    config = data.get("config", {})
    from backend.emailer import format_daily_report, format_html_report, send_email_report
    report_text = format_daily_report(products, transactions)
    html_body = format_html_report(products, transactions)
    result = send_email_report(config, report_text, html_body)
    if result.get("success"):
        logger.info("Daily test email sent successfully")
        return {"success": True, "message": "Email test report sent successfully!", "data": result}
    else:
        logger.error(f"Daily test email failed: {result.get('error')}")
        raise HTTPException(status_code=400, detail=f"Failed to send: {result.get('error')}")

@app.post("/api/email/test-send-weekly")
def trigger_weekly_email_test_send(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    transactions = data.get("transactions", [])
    config = data.get("config", {})
    from backend.emailer import format_weekly_report, format_weekly_html_report, send_email_report
    report_text, period = format_weekly_report(products, transactions)
    html_body = format_weekly_html_report(products, transactions)
    subject = f"Solar Store Weekly Report - {period}"
    result = send_email_report(config, report_text, html_body, subject=subject)
    if result.get("success"):
        logger.info("Weekly test email sent successfully")
        return {"success": True, "message": "Weekly email test report sent successfully!", "data": result}
    else:
        logger.error(f"Weekly test email failed: {result.get('error')}")
        raise HTTPException(status_code=400, detail=f"Failed to send: {result.get('error')}")

@app.get("/api/whatsapp/config")
def get_whatsapp_config(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    return {
        "whatsapp_recipient": config.get("whatsapp_recipient", ""),
        "whatsapp_phone_number_id": config.get("whatsapp_phone_number_id", ""),
        "whatsapp_token": config.get("whatsapp_token", "")
    }

@app.post("/api/whatsapp/config")
def save_whatsapp_config(payload: WhatsAppConfigModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    config["whatsapp_recipient"] = payload.whatsapp_recipient
    config["whatsapp_phone_number_id"] = payload.whatsapp_phone_number_id
    config["whatsapp_token"] = payload.whatsapp_token
    data["config"] = config
    write_json(STOCK_FILE_PATH, data)
    logger.info("WhatsApp configuration updated")
    return {"success": True, "message": "WhatsApp settings saved successfully"}

class EnergyPeakSunModel(BaseModel):
    energy_peak_sun_hours: float = Field(default=5.0, ge=1, le=14)

@app.get("/api/energy/config")
def get_energy_config(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    return {
        "energy_peak_sun_hours": float(config.get("energy_peak_sun_hours", 5))
    }

@app.post("/api/energy/config")
def save_energy_config(payload: EnergyPeakSunModel, token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    config["energy_peak_sun_hours"] = payload.energy_peak_sun_hours
    data["config"] = config
    write_json(STOCK_FILE_PATH, data)
    logger.info("Energy forecast settings updated")
    return {"success": True, "message": "Energy forecast settings saved"}

@app.post("/api/whatsapp/test-send")
def trigger_whatsapp_test_send(token: str = Depends(verify_token)):
    data = read_json(STOCK_FILE_PATH)
    config = data.get("config", {})
    phone_number_id = config.get("whatsapp_phone_number_id", "")
    whatsapp_token = config.get("whatsapp_token", "")
    recipient = config.get("whatsapp_recipient", "")
    if not phone_number_id or not whatsapp_token or not recipient:
        raise HTTPException(status_code=400, detail="WhatsApp not fully configured. Save Phone Number ID, Token, and Recipient first.")
    from backend.whatsapp import send_test_message
    result = send_test_message(phone_number_id, whatsapp_token, recipient)
    if result.get("success"):
        logger.info("WhatsApp test message sent successfully")
        return {"success": True, "message": "WhatsApp test message sent!", "data": result}
    else:
        logger.error(f"WhatsApp test failed: {result.get('error')}")
        raise HTTPException(status_code=400, detail=f"WhatsApp send failed: {result.get('error')}")

@app.get("/api/reports/export")
def export_report(
    range_type: str = Query("daily", pattern="^(daily|weekly|monthly)$"),
    file_format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    token: str = Depends(verify_token)
):
    data = read_json(STOCK_FILE_PATH)
    products = data.get("products", [])
    transactions = data.get("transactions", [])
    today = datetime.date.today()
    if range_type == "daily":
        start_date = today
        end_date = today
    elif range_type == "weekly":
        start_date = today - datetime.timedelta(days=7)
        end_date = today
    elif range_type == "monthly":
        start_date = today - datetime.timedelta(days=30)
        end_date = today
    filtered_txs = []
    for tx in transactions:
        tx_date_str = tx.get("date", "")
        try:
            tx_date = datetime.datetime.strptime(tx_date_str, "%Y-%m-%d").date()
            if start_date <= tx_date <= end_date:
                filtered_txs.append(tx)
        except ValueError:
            continue
    filtered_txs = sorted(filtered_txs, key=lambda x: x.get("timestamp", ""))
    timestamp_str = today.strftime("%Y-%m-%d")
    if file_format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Report Summary"
        ws_summary.views.sheetView[0].showGridLines = True
        title_font = openpyxl.styles.Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_font = openpyxl.styles.Font(name="Calibri", size=12, bold=True)
        bold_font = openpyxl.styles.Font(name="Calibri", size=10, bold=True)
        regular_font = openpyxl.styles.Font(name="Calibri", size=10)
        header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        accent_fill = openpyxl.styles.PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        zebra_fill = openpyxl.styles.PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='DDDDDD'),
            right=openpyxl.styles.Side(style='thin', color='DDDDDD'),
            top=openpyxl.styles.Side(style='thin', color='DDDDDD'),
            bottom=openpyxl.styles.Side(style='thin', color='DDDDDD')
        )
        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = f"ENERGY SOLAR STOCK REPORT - {range_type.upper()}"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].fill = header_fill
        ws_summary["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_summary.row_dimensions[1].height = 40
        ws_summary["A3"] = "Report Date:"
        ws_summary["A3"].font = bold_font
        ws_summary["B3"] = today.strftime("%d/%m/%Y")
        ws_summary["B3"].font = regular_font
        ws_summary["A4"] = "Period:"
        ws_summary["A4"].font = bold_font
        ws_summary["B4"] = f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
        ws_summary["B4"].font = regular_font
        ws_summary["A6"] = "Key Stock Indicators"
        ws_summary["A6"].font = section_font
        ws_summary["A7"] = "Indicator"
        ws_summary["B7"] = "Value"
        for cell in ["A7", "B7"]:
            ws_summary[cell].font = header_font
            ws_summary[cell].fill = accent_fill
        indicators = [
            ("Total Products in System", len(products)),
            ("Total Current Available Qty", sum(p.get("quantity", 0) for p in products)),
            ("Low Stock Alert Items", sum(1 for p in products if p.get("quantity", 0) < p.get("minimum_stock", 0))),
            ("Out of Stock Items", sum(1 for p in products if p.get("quantity", 0) == 0)),
            ("Total Inflow Qty (in Period)", sum(tx.get("quantity", 0) for tx in filtered_txs if tx.get("type") == "IN")),
            ("Total Outflow Qty (in Period)", sum(tx.get("quantity", 0) for tx in filtered_txs if tx.get("type") == "OUT"))
        ]
        for r_idx, (ind, val) in enumerate(indicators, start=8):
            ws_summary.cell(row=r_idx, column=1, value=ind).font = regular_font
            ws_summary.cell(row=r_idx, column=2, value=val).font = bold_font
            ws_summary.cell(row=r_idx, column=1).border = thin_border
            ws_summary.cell(row=r_idx, column=2).border = thin_border
            if r_idx % 2 == 1:
                ws_summary.cell(row=r_idx, column=1).fill = zebra_fill
                ws_summary.cell(row=r_idx, column=2).fill = zebra_fill
        ws_prod = wb.create_sheet(title="Current Catalog")
        ws_prod.views.sheetView[0].showGridLines = True
        headers_prod = ["Product ID", "Name", "Category", "Brand", "Unit", "Qty In Stock", "Min Qty", "Rack Location", "Supplier"]
        ws_prod.append(headers_prod)
        for col_idx, text in enumerate(headers_prod, start=1):
            cell = ws_prod.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        for r_idx, p in enumerate(products, start=2):
            vals = [p.get("id"), p.get("name"), p.get("category"), p.get("brand"), p.get("unit"), p.get("quantity"), p.get("minimum_stock"), p.get("rack_location"), p.get("supplier")]
            ws_prod.append(vals)
            for c_idx in range(1, len(vals) + 1):
                cell = ws_prod.cell(row=r_idx, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border
                if c_idx == 6:
                    qty = p.get("quantity", 0)
                    min_s = p.get("minimum_stock", 0)
                    if qty == 0:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = openpyxl.styles.Font(name="Calibri", size=10, bold=True, color="991B1B")
                    elif qty < min_s:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = openpyxl.styles.Font(name="Calibri", size=10, bold=True, color="92400E")
                elif r_idx % 2 == 1:
                    cell.fill = zebra_fill
        ws_tx = wb.create_sheet(title="Transactions History")
        ws_tx.views.sheetView[0].showGridLines = True
        headers_tx = ["TX ID", "Date", "Type", "Product ID", "Product Name", "Quantity", "Customer/Site", "Employee/Supplier", "Remarks"]
        ws_tx.append(headers_tx)
        for col_idx, text in enumerate(headers_tx, start=1):
            cell = ws_tx.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        for r_idx, tx in enumerate(filtered_txs, start=2):
            tx_type = tx.get("type", "IN")
            if tx_type == "IN":
                cust_site = "-"
                emp_supp = tx.get("entity") or "-"
            else:
                cust_site = tx.get("entity") or "-"
                emp_supp = tx.get("employee") or "-"
            vals = [tx.get("id"), tx.get("date"), tx.get("type"), tx.get("product_id"), tx.get("product_name"), tx.get("quantity"), cust_site, emp_supp, tx.get("remarks")]
            ws_tx.append(vals)
            for c_idx in range(1, len(vals) + 1):
                cell = ws_tx.cell(row=r_idx, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border
                if c_idx == 3:
                    txtype = tx.get("type")
                    if txtype == "IN":
                        cell.fill = openpyxl.styles.PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = openpyxl.styles.Font(name="Calibri", size=10, bold=True, color="065F46")
                    else:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = openpyxl.styles.Font(name="Calibri", size=10, bold=True, color="991B1B")
                elif r_idx % 2 == 1:
                    cell.fill = zebra_fill
        for ws in [ws_summary, ws_prod, ws_tx]:
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        if ws.title == "Report Summary" and cell.row == 1:
                            continue
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        filename = f"{range_type}_report_{timestamp_str}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        pdf_stream = io.BytesIO()
        doc = SimpleDocTemplate(pdf_stream, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=22, textColor=colors.HexColor('#1E293B'), spaceAfter=6)
        subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=11, textColor=colors.HexColor('#F59E0B'), spaceAfter=15)
        section_style = ParagraphStyle('ReportSection', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.HexColor('#1E293B'), spaceBefore=15, spaceAfter=8)
        meta_label_style = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#64748B'))
        meta_val_style = ParagraphStyle('MetaValue', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#0F172A'))
        table_hdr_style = ParagraphStyle('TableHdr', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=1)
        table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#334155'))
        table_cell_bold_style = ParagraphStyle('TableCellBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#0F172A'))
        story.append(Paragraph("Solar Stock Management System", title_style))
        story.append(Paragraph(f"{range_type.upper()} TRANSACTION & CATALOG REPORT", subtitle_style))
        story.append(Spacer(1, 10))
        meta_data = [
            [Paragraph("Report Period:", meta_label_style), Paragraph(f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}", meta_val_style), Paragraph("Export Date:", meta_label_style), Paragraph(today.strftime('%d/%m/%Y'), meta_val_style)],
            [Paragraph("Total Cataloged:", meta_label_style), Paragraph(f"{len(products)} products", meta_val_style), Paragraph("Total Current Stock:", meta_label_style), Paragraph(f"{sum(p.get('quantity',0) for p in products)} items", meta_val_style)]
        ]
        meta_table = Table(meta_data, colWidths=[100, 160, 100, 160])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor('#F1F5F9')),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 15))
        low_stock_p = [p for p in products if p.get("quantity", 0) < p.get("minimum_stock", 0)]
        if low_stock_p:
            story.append(Paragraph("Low Stock & Out of Stock Warnings", section_style))
            low_headers = ["ID", "Product Name", "Category", "Qty In Stock", "Min Stock", "Rack Location"]
            low_table_data = [[Paragraph(h, table_hdr_style) for h in low_headers]]
            for p in low_stock_p:
                low_table_data.append([
                    Paragraph(p.get("id", ""), table_cell_bold_style),
                    Paragraph(p.get("name", ""), table_cell_style),
                    Paragraph(p.get("category", ""), table_cell_style),
                    Paragraph(str(p.get("quantity", 0)), table_cell_bold_style),
                    Paragraph(str(p.get("minimum_stock", 0)), table_cell_style),
                    Paragraph(p.get("rack_location", ""), table_cell_style)
                ])
            low_table = Table(low_table_data, colWidths=[50, 150, 90, 80, 80, 90])
            low_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#DC2626')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFF5F5'), colors.white]),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#FCA5A5')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#FEE2E2')),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(low_table)
            story.append(Spacer(1, 15))
        story.append(Paragraph(f"Transaction Records ({range_type.capitalize()})", section_style))
        if not filtered_txs:
            story.append(Paragraph("No transaction records found for this period.", table_cell_style))
        else:
            tx_headers = ["TX ID", "Date", "Type", "Product Name", "Qty", "Customer/Site", "Employee/Supplier", "Remarks"]
            tx_table_data = [[Paragraph(h, table_hdr_style) for h in tx_headers]]
            for tx in filtered_txs:
                t_color = '#065F46' if tx.get("type") == "IN" else '#991B1B'
                tx_type_para = Paragraph(f"<font color='{t_color}'><b>{tx.get('type')}</b></font>", table_cell_bold_style)
                tx_type = tx.get("type", "IN")
                if tx_type == "IN":
                    cust_site = "-"
                    emp_supp = tx.get("entity") or "-"
                else:
                    cust_site = tx.get("entity") or "-"
                    emp_supp = tx.get("employee") or "-"
                tx_table_data.append([
                    Paragraph(tx.get("id", ""), table_cell_bold_style),
                    Paragraph(tx.get("date", ""), table_cell_style),
                    tx_type_para,
                    Paragraph(tx.get("product_name", ""), table_cell_style),
                    Paragraph(str(tx.get("quantity", 0)), table_cell_bold_style),
                    Paragraph(cust_site, table_cell_style),
                    Paragraph(emp_supp, table_cell_style),
                    Paragraph(tx.get("remarks") or "-", table_cell_style)
                ])
            tx_table = Table(tx_table_data, colWidths=[45, 55, 40, 115, 35, 90, 80, 80])
            tx_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F8FAFC'), colors.white]),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(tx_table)
        doc.build(story)
        pdf_stream.seek(0)
        filename = f"{range_type}_report_{timestamp_str}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        with open(filepath, "wb") as f_out:
            f_out.write(pdf_stream.getbuffer())
        pdf_stream.seek(0)
        return StreamingResponse(
            pdf_stream,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@app.get("/api/reports/list")
def list_reports(token: str = Depends(verify_token)):
    try:
        files = os.listdir(REPORTS_DIR)
        reports = []
        for f in files:
            fpath = os.path.join(REPORTS_DIR, f)
            if os.path.isfile(fpath):
                reports.append({
                    "name": f,
                    "size": os.path.getsize(fpath),
                    "modified": datetime.datetime.fromtimestamp(os.path.getmtime(fpath)).isoformat()
                })
        return sorted(reports, key=lambda r: r["name"], reverse=True)
    except Exception as e:
        logger.error(f"Failed to list reports: {e}")
        return []

app.mount("/css", StaticFiles(directory=os.path.join(BASE_DIR, "css"), html=True), name="css")
app.mount("/js", StaticFiles(directory=os.path.join(BASE_DIR, "js"), html=True), name="js")

@app.get("/")
def get_index():
    return FileResponse(os.path.join(BASE_DIR, "index.html"))

@app.exception_handler(404)
async def custom_404(request, exc):
    return FileResponse(os.path.join(BASE_DIR, "index.html"))

if __name__ == "__main__":
    import uvicorn
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    reload_val = os.environ.get("RELOAD", "false").lower() == "true"
    logger.info(f"Starting server on {host}:{port} (reload={reload_val})")
    uvicorn.run("backend.app:app", host=host, port=port, reload=reload_val)
